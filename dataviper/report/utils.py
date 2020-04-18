from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, colors
from openpyxl.utils.dataframe import dataframe_to_rows


def create_workbook_from_dataframe(df):
    """
    1. Create workbook from specified pandas.DataFrame
    2. Adjust columns width to fit the text inside
    3. Make the index column and the header row bold
    4. Fill background color for the header row

    Other beautification MUST be done by usage side.
    """
    workbook = Workbook()
    ws = workbook.active

    rows = dataframe_to_rows(df.reset_index(), index=False)
    col_widths = [0] * (len(df.columns) + 1)
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):

            if type(val) is str:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            elif hasattr(val, "sort"):
                cell = ws.cell(row=i, column=j, value=", ".join(list(map(lambda v: str(v), list(val)))))
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            else:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val)) + 1])

            # Make the index column and the header row bold
            if i == 1 or j == 1:
                cell.font = Font(bold=True)

            # Fill background color for the header row
            if i == 1:
                cell.fill = PatternFill('solid', fgColor=colors.YELLOW)

    # Adjust column width
    for i, w in enumerate(col_widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = w

    return workbook


def fill_cell_color(cell, color, fill_type='solid'):
    cell.fill = PatternFill(fill_type, fgColor=('00' + color))


def fill_text_color(cell, color, bold=False):
    cell.font = Font(color=color, bold=bold)
